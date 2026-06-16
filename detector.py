"""Core detection engine."""

from __future__ import annotations

import gzip
import io
import zipfile
from dataclasses import dataclass, field

from secureproxy.config import Config
from secureproxy.evasion import DecodedBlob, expand_text_for_scanning
from secureproxy.luhn import luhn_check
from secureproxy.patterns import EVASION_PATTERNS, PATTERNS
from secureproxy.parser import ParsedContent, parse_flow_content


@dataclass
class ScanResult:
    rule_hits: dict[str, list[tuple[str, str]]] = field(default_factory=dict)
    evasion_blobs: list[DecodedBlob] = field(default_factory=list)
    triggered_rules: set[str] = field(default_factory=set)
    bytes_scanned: int = 0
    truncated: bool = False


class DetectionEngine:
    def __init__(self, config: Config) -> None:
        self.config = config
        self.body_limit = config.body_limits.get("default_max", 1_048_576)
        self.scan_max = config.body_limits.get("scan_max", 10_485_760)

    def scan_text(self, source: str, text: str, result: ScanResult) -> None:
        if not text:
            return

        expanded, blobs = expand_text_for_scanning(text)
        result.evasion_blobs.extend(blobs)

        for blob in blobs:
            self._run_patterns(f"{source}:decoded_{blob.encoding}", blob.decoded, result)

        self._run_patterns(source, expanded, result)

        for evasion_rule, pattern in EVASION_PATTERNS.items():
            if pattern.search(text) and self._rule_enabled(evasion_rule):
                self._add_hit(result, evasion_rule, source, text[:80])

    def _run_patterns(self, source: str, text: str, result: ScanResult) -> None:
        for spec in PATTERNS:
            if not self._rule_enabled(spec.rule):
                continue
            if spec.validator == "luhn":
                for match in spec.pattern.finditer(text):
                    if luhn_check(match.group()):
                        self._add_hit(result, spec.rule, source, match.group())
            elif spec.validator == "bip39_word_count":
                for match in spec.pattern.finditer(text):
                    words = match.group().split()
                    if 12 <= len(words) <= 24 and len(words) % 3 == 0:
                        self._add_hit(result, spec.rule, source, match.group()[:80])
            else:
                match = spec.pattern.search(text)
                if match:
                    self._add_hit(result, spec.rule, source, match.group())

    def _add_hit(
        self, result: ScanResult, rule: str, source: str, sample: str
    ) -> None:
        result.triggered_rules.add(rule)
        result.rule_hits.setdefault(rule, []).append((source, sample))

    def _rule_enabled(self, rule: str) -> bool:
        cfg = self.config.rules.get(rule)
        return cfg.enabled if cfg else False

    def scan_parsed_content(self, parsed: ParsedContent, result: ScanResult) -> None:
        for source, text in parsed.segments:
            self.scan_text(source, text, result)

    def scan_body_bytes(
        self,
        body_bytes: bytes,
        content_type: str,
        result: ScanResult,
        file_inspector: "FileInspector | None" = None,
    ) -> None:
        if not body_bytes:
            return

        if len(body_bytes) > self.scan_max:
            body_bytes = body_bytes[: self.scan_max]
            result.truncated = True

        result.bytes_scanned += len(body_bytes)

        if "gzip" in content_type or body_bytes[:2] == b"\x1f\x8b":
            try:
                decompressed = gzip.decompress(body_bytes)
                text = decompressed.decode("utf-8", errors="ignore")
                self.scan_text("gzip_body", text, result)
                if file_inspector:
                    file_inspector.inspect_bytes(decompressed, "decompressed.gz", result)
                return
            except Exception:
                pass

        text = body_bytes.decode("utf-8", errors="ignore")
        if text.strip():
            self.scan_text("body", text, result)

        if file_inspector:
            file_inspector.inspect_bytes(body_bytes, content_type, result)


class FileInspector:
    TEXT_EXTENSIONS = {".txt", ".csv", ".json", ".env", ".log", ".md", ".xml", ".yaml", ".yml"}
    ARCHIVE_EXTENSIONS = {".zip"}

    def __init__(self, engine: DetectionEngine) -> None:
        self.engine = engine

    def inspect_bytes(self, data: bytes, hint: str, result: ScanResult) -> None:
        hint_lower = hint.lower()

        if any(hint_lower.endswith(ext) for ext in self.ARCHIVE_EXTENSIONS) or data[:2] == b"PK":
            self._inspect_zip(data, result)
            return

        if any(hint_lower.endswith(ext) for ext in self.TEXT_EXTENSIONS) or "json" in hint_lower:
            text = data.decode("utf-8", errors="ignore")
            self.engine.scan_text(f"file:{hint}", text, result)
            return

        if hint_lower.endswith(".pdf"):
            self._inspect_pdf(data, result)
        elif hint_lower.endswith(".docx"):
            self._inspect_docx(data, result)
        elif hint_lower.endswith(".xlsx"):
            self._inspect_xlsx(data, result)

    def _inspect_zip(self, data: bytes, result: ScanResult) -> None:
        try:
            with zipfile.ZipFile(io.BytesIO(data)) as zf:
                for info in zf.infolist():
                    if info.is_dir() or info.file_size > self.engine.body_limit:
                        continue
                    try:
                        content = zf.read(info.filename)
                        self.inspect_bytes(content, info.filename, result)
                    except Exception:
                        continue
        except zipfile.BadZipFile:
            pass

    def _inspect_pdf(self, data: bytes, result: ScanResult) -> None:
        try:
            from pypdf import PdfReader

            reader = PdfReader(io.BytesIO(data))
            text_parts = [page.extract_text() or "" for page in reader.pages[:20]]
            self.engine.scan_text("file:pdf", "\n".join(text_parts), result)
        except Exception:
            pass

    def _inspect_docx(self, data: bytes, result: ScanResult) -> None:
        try:
            from docx import Document

            doc = Document(io.BytesIO(data))
            text = "\n".join(p.text for p in doc.paragraphs)
            self.engine.scan_text("file:docx", text, result)
        except Exception:
            pass

    def _inspect_xlsx(self, data: bytes, result: ScanResult) -> None:
        try:
            from openpyxl import load_workbook

            wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            rows: list[str] = []
            for sheet in wb.worksheets:
                for row in sheet.iter_rows(max_row=500, values_only=True):
                    rows.append(" ".join(str(c) for c in row if c is not None))
            self.engine.scan_text("file:xlsx", "\n".join(rows), result)
        except Exception:
            pass
